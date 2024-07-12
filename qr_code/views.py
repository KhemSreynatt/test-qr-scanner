
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
import json
import qrcode
from django.http import HttpResponse
from django.utils import timezone
from io import BytesIO
from django.shortcuts import render
from openpyxl import load_workbook


@csrf_exempt
def qr_result(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        result = data.get('result')
        # Process the result as needed
        return JsonResponse({'status': 'success', 'message': f'Processed QR: {result}'})
    return JsonResponse({'status': 'error'}, status=400)



def generate_qr(request, qrcodes):
    event_id = f"{qrcodes}_{timezone.now().strftime('%Y%m%d%H%M%S')}"
    qr_data = json.dumps({
        "qrcodes": qrcodes,
    })
    img = qrcode.make(qrcodes)
    response = HttpResponse(content_type="image/png")
    buffer = BytesIO()
    img.save(buffer, "PNG")
    response.write(buffer.getvalue())
    return response

@csrf_exempt
def handle_scan(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        name = data.get('name')
        id = data.get('id')
        address = data.get('address')
        # Path to your Excel file
        excel_file = 'assets/data_attendance.xlsx'
        try:
            workbook = load_workbook(excel_file)
            sheet = workbook.active
            next_row = sheet.max_row + 1
            sheet.cell(row=next_row, column=1, value=name)
            sheet.cell(row=next_row, column=2, value=id)
            sheet.cell(row=next_row, column=3, value=address)
            workbook.save(excel_file)
            return JsonResponse({'status': 'success'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'})

# def index(request):
#     return render(request, 'assets/data_attendance.xlsx')